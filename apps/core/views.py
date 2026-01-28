from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from apps.vehicles.models import VehicleData, EmissionFactor


def login_view(request):
    """Vue de connexion"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            auth_login(request, user)
            next_url = request.POST.get('next') or request.GET.get('next') or 'dashboard'
            return redirect(next_url)
        else:
            messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
    
    return render(request, 'core/login.html')


def logout_view(request):
    """Vue de déconnexion"""
    auth_logout(request)
    messages.success(request, 'Vous avez été déconnecté avec succès.')
    return redirect('login')


@login_required
def dashboard_view(request):
    """Vue du tableau de bord"""
    from django.db.models import Sum
    from datetime import datetime
    from apps.vehicles.models import VehicleData
    from apps.purchases.models import PurchaseData
    from apps.alimentation.models import FoodEntry
    from apps.batiment.models import BuildingEnergyData
    from apps.numerique.models import EquipementNumerique
    from apps.sensibilisation.services import SensibilisationService
    from apps.sensibilisation.models import MessageSensibilisation

    current_year = datetime.now().year

    # Le dashboard est global pour tous les utilisateurs connectés.
    # On récupère toutes les données, sans filtrer.
    base_queryset_vehicles = VehicleData.objects.all()
    base_queryset_purchases = PurchaseData.objects.all()
    base_queryset_alimentation = FoodEntry.objects.all()
    base_queryset_batiment = BuildingEnergyData.objects.all()
    base_queryset_numerique = EquipementNumerique.objects.all()


    # Calcul des totaux pour l'année en cours
    vehicles_total = float(base_queryset_vehicles.filter(year=current_year).aggregate(total=Sum('total_co2_kg'))['total'] or 0)
    purchases_total = float(base_queryset_purchases.filter(year=current_year).aggregate(total=Sum('total_co2_kg'))['total'] or 0)
    alimentation_total = float(base_queryset_alimentation.filter(year=current_year).aggregate(total=Sum('total_co2_kg'))['total'] or 0)
    building_total = float(base_queryset_batiment.filter(year=current_year).aggregate(total=Sum('total_co2_kg'))['total'] or 0)
    numerique_total = float(base_queryset_numerique.filter(year=current_year).aggregate(total=Sum('total_co2_kg'))['total'] or 0)

    total_co2 = vehicles_total + purchases_total + alimentation_total + building_total + numerique_total

    # Sensibilisation
    stats_modules = {
        'total': total_co2,
        'vehicles': vehicles_total,
        'numerique': numerique_total,
        'batiment': building_total
    }
    
    context = {
        'vehicle_count': base_queryset_vehicles.count(),
        'emission_factors_count': EmissionFactor.objects.filter(is_active=True).count(),
        # Sensibilisation
        'message_admin': MessageSensibilisation.objects.filter(actif=True).first(),
        'equivalences': SensibilisationService.get_equivalences(total_co2),
        'conseils': SensibilisationService.get_conseils_automatiques(stats_modules),
        'total_co2': total_co2
    }
    return render(request, 'core/dashboard.html', context)


@login_required
def send_reminder_email(request):
    """
    Envoie un email de rappel à tous les agents pour saisir leurs données.
    Réservé aux administrateurs.
    """
    from django.contrib.auth.models import User
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.conf import settings
    from django.utils import timezone
    import logging
    
    # Vérifier que l'utilisateur est admin
    if not request.user.is_staff:
        messages.error(request, "⛔ Accès refusé. Cette fonctionnalité est réservée aux administrateurs.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        # Récupérer tous les users actifs (sauf admins et ceux sans email)
        recipients = User.objects.filter(
            is_active=True,
            is_staff=False
        ).exclude(email='')
        
        # Envoyer l'email
        sent_count = 0
        failed_count = 0
        
        for user in recipients:
            try:
                # Préparer le message
                message = render_to_string('emails/reminder.txt', {
                    'user': user,
                    'year': timezone.now().year,
                    'domain': request.get_host()
                })
                
                # Envoyer
                send_mail(
                    subject=f"Rappel - Saisie du bilan carbone {timezone.now().year}",
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[user.email],
                    fail_silently=False,
                )
                sent_count += 1
                
            except Exception as e:
                failed_count += 1
                logging.error(f"Erreur envoi email à {user.email}: {e}")
        
        # Message de retour
        if sent_count > 0:
            messages.success(request, f"✅ {sent_count} email(s) envoyé(s) avec succès")
        if failed_count > 0:
            messages.warning(request, f"⚠️ {failed_count} email(s) n'ont pas pu être envoyé(s)")
        
        return redirect('dashboard')
    
    # GET : Afficher page de confirmation
    recipients_count = User.objects.filter(
        is_active=True,
        is_staff=False
    ).exclude(email='').count()
    
    return render(request, 'core/confirm_send_email.html', {
        'recipients_count': recipients_count,
        'current_year': timezone.now().year
    })


@login_required
def dashboard_emissions_api(request):
    """
    API endpoint pour récupérer les données d'émissions agrégées
    pour le graphique de répartition globale du dashboard.
    
    Retourne un JSON avec les émissions par secteur.
    """
    from django.http import JsonResponse
    from django.db.models import Sum
    from apps.purchases.models import PurchaseData
    from apps.alimentation.models import FoodEntry
    from apps.batiment.models import BuildingEnergyData
    from apps.numerique.models import EquipementNumerique
    
    from django.utils import timezone
    
    current_year = timezone.now().year
    
    # Agréger les données par secteur pour TOUS les utilisateurs (vue globale admin) pour l'ANNEE EN COURS
    vehicles_total = float(VehicleData.objects.filter(year=current_year).aggregate(total=Sum('total_co2_kg'))['total'] or 0)
    
    purchases_total = float(PurchaseData.objects.filter(year=current_year).aggregate(total=Sum('total_co2_kg'))['total'] or 0)
    
    alimentation_total = float(FoodEntry.objects.filter(year=current_year).aggregate(total=Sum('total_co2_kg'))['total'] or 0)
    
    building_total = float(BuildingEnergyData.objects.filter(year=current_year).aggregate(total=Sum('total_co2_kg'))['total'] or 0)

    numerique_total = float(EquipementNumerique.objects.filter(year=current_year).aggregate(total=Sum('total_co2_kg'))['total'] or 0)
    
    # Préparer les données pour Chart.js
    data = {
        'labels': ['Véhicules', 'Achats', 'Alimentation', 'Bâtiments', 'Numérique'],
        'data': [
            round(vehicles_total, 2),
            round(purchases_total, 2),
            round(alimentation_total, 2),
            round(building_total, 2),
            round(numerique_total, 2)
        ],
        'colors': ['#4A90E2', '#9B59B6', '#27AE60', '#F39C12', '#34495E'],
        'total': round(vehicles_total + purchases_total + alimentation_total + building_total + numerique_total, 2)
    }
    
    return JsonResponse(data)


@login_required
def export_data_view(request):
    """Vue pour exporter les données (CSV ou Excel)"""
    import csv
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import DoughnutChart, Reference, Series
    from openpyxl.chart.label import DataLabelList
    
    from apps.vehicles.models import VehicleData
    from apps.batiment.models import BuildingEnergyData
    
    format_type = request.GET.get('format', 'csv')
    
    # Admins exportent tout, les agents exportent les données de leur groupe
    if request.user.is_staff or request.user.is_superuser:
        vehicles = VehicleData.objects.all()
        buildings = BuildingEnergyData.objects.all()
    else:
        user_groups = request.user.groups.all()
        vehicles = VehicleData.objects.filter(group__in=user_groups)
        buildings = BuildingEnergyData.objects.filter(group__in=user_groups)
    
    if format_type == 'xlsx':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="bilan_carbone_evry.xlsx"'
        
        wb = Workbook()
        
        # --- Feuille 1: Synthèse ---
        ws_synth = wb.active
        ws_synth.title = "Synthèse"
        
        # Calcul des totaux
        total_batiment = sum(float(b.total_co2_kg or 0) for b in buildings)
        total_vehicle = sum(float(v.total_co2_kg or 0) for v in vehicles)
        total_global = total_batiment + total_vehicle
        
        ws_synth.append(['Catégorie', 'Emissions (kgCO2e)'])
        ws_synth.append(['Bâtiments', total_batiment])
        ws_synth.append(['Véhicules', total_vehicle])
        
        # Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2d6a4f", end_color="2d6a4f", fill_type="solid")
        
        for cell in ws_synth[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            
        # Graphique
        chart = DoughnutChart()
        chart.title = "Répartition Bilan Carbone"
        chart.style = 10
        chart.holeSize = 70
        
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        
        cats = Reference(ws_synth, min_col=1, min_row=2, max_row=3)
        data = Reference(ws_synth, min_col=2, min_row=1, max_row=3)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        ws_synth.add_chart(chart, "D2")
        
        # --- Feuille 2: Détails ---
        ws_detail = wb.create_sheet(title="Détails")
        ws_detail.append(['Type', 'Année', 'Nom/Service', 'Total CO2 (kg)'])
        
        for b in buildings:
            ws_detail.append(['Bâtiment', b.year, b.site_name, b.total_co2_kg])
            
        for v in vehicles:
            ws_detail.append(['Véhicule', v.year, v.service, v.total_co2_kg])

        wb.save(response)
        return response
        
    else: # CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="bilan_carbone_evry.csv"'
        
        writer = csv.writer(response)
        
        # Totaux
        total_batiment = sum(float(b.total_co2_kg or 0) for b in buildings)
        total_vehicle = sum(float(v.total_co2_kg or 0) for v in vehicles)
        total_global = total_batiment + total_vehicle
        
        pct_batimes = int((total_batiment/total_global)*100) if total_global else 0
        pct_vehicles = int((total_vehicle/total_global)*100) if total_global else 0
        
        writer.writerow(['BILAN CARBONE - SYNTHÈSE'])
        writer.writerow(['Total Global', f'{total_global:.2f} kgCO2e'])
        writer.writerow(['Bâtiments', f'{total_batiment:.2f} kgCO2e', f'{pct_batimes}%'])
        writer.writerow(['Véhicules', f'{total_vehicle:.2f} kgCO2e', f'{pct_vehicles}%'])
        writer.writerow([])
        
        writer.writerow(['DETAILS'])
        writer.writerow(['Type', 'Année', 'Nom/Service', 'Total CO2 (kg)'])
        
        for b in buildings:
            writer.writerow(['Bâtiment', b.year, b.site_name, b.total_co2_kg])
            
        for v in vehicles:
            writer.writerow(['Véhicule', v.year, v.service, v.total_co2_kg])
            
        return response


@login_required
def statistics_view(request):
    """
    Vue détaillée pour les statistiques et comparaisons annuelles.
    """
    from django.db.models import Sum
    from django.utils import timezone
    from apps.vehicles.models import VehicleData
    from apps.purchases.models import PurchaseData
    from apps.alimentation.models import FoodEntry
    from apps.batiment.models import BuildingEnergyData
    
    context = {}
    return render(request, 'core/statistics.html', context)


@login_required
def statistics_api(request):
    """
    API pour les statistiques:
    - Top Chart: Stacked Bar (Global par secteur)
    - Bottom Charts: Détails par secteur (Doughnuts mix énergétique, mix véhicules...)
    """
    from django.http import JsonResponse
    from django.db.models import Sum
    from django.utils import timezone
    from apps.vehicles.models import VehicleData
    from apps.purchases.models import PurchaseData
    from apps.alimentation.models import FoodEntry, FoodEmissionFactor
    from apps.batiment.models import BuildingEnergyData
    from apps.numerique.models import EquipementNumerique
    
    # 1. Identifier les années disponibles (vue globale)
    years = set()
    years.update(VehicleData.objects.values_list('year', flat=True))
    years.update(BuildingEnergyData.objects.values_list('year', flat=True))
    years.update(FoodEntry.objects.values_list('year', flat=True))
    years.update(PurchaseData.objects.values_list('year', flat=True))
    years.update(EquipementNumerique.objects.values_list('year', flat=True))
    
    sorted_years = sorted(list(years))
    if not sorted_years:
        return JsonResponse({'years': [], 'stacked_data': {}, 'details_data': {}})
        sorted_years = [timezone.now().year]

    # 2. Préparer les données pour le Stacked Bar Chart (Global)
    # 4 arrays alignés sur sorted_years
    stacked_data = {
        'vehicles': [],
        'buildings': [],
        'food': [],
        'purchases': [],
        'numerique': []
    }
    
    # 3. Préparer les détails pour les Doughnuts (Dict {year: {sector: {label: value}}})
    details_data = {}

    # Pré-chargement des facteurs Alimentation pour éviter N requêtes
    food_factors_map = {f.code: f.kg_co2_per_meal for f in FoodEmissionFactor.objects.all()}
    
    for year in sorted_years:
        details_data[year] = {
            'buildings': {},
            'vehicles': {},
            'food': {},
            'purchases': {},
            'numerique': {}
        }

        # --- BUILDING DETAILS ---
        qs_b = BuildingEnergyData.objects.filter(year=year)
        # On recalcule les sommes des 'parties' pour le camembert
        # Attention: BuildingEnergyData n'a pas les champs 'elec_co2' stockés, on doit calculer
        # Sum( consumption * factor ) -> comme factor est par ligne, c'est complexe si factor change
        # Simplification: On itère pour sommer (pas optimal si 1M lignes mais ok ici)
        b_elec = 0
        b_gas = 0
        b_heat = 0
        b_cool = 0
        total_b = 0
        
        for b in qs_b:
            b_elec += float(b.electricity_kwh * b.electricity_factor)
            b_gas += float(b.gas_kwh * b.gas_factor)
            b_heat += float(b.heating_network_kwh * b.heating_network_factor)
            b_cool += float(b.cooling_kwh * b.cooling_factor)
            total_b += float(b.total_co2_kg)
            
        details_data[year]['buildings'] = {
            'Électricité': b_elec,
            'Gaz Naturel': b_gas,
            'Réseau Chaleur': b_heat,
            'Climatisation': b_cool
        }
        stacked_data['buildings'].append(total_b)


        # --- VEHICLE DETAILS ---
        qs_v = VehicleData.objects.filter(year=year)
        v_essence = 0
        v_gazole = 0
        v_distance = 0
        total_v = 0
        
        for v in qs_v:
            total_v += float(v.total_co2_kg or 0)
            if v.calculation_method == 'fuel':
                v_essence += float(v.essence_co2_kg or 0)
                v_gazole += float(v.gazole_co2_kg or 0)
            else:
                v_distance += float(v.total_co2_kg or 0)
                
        details_data[year]['vehicles'] = {
            'Essence': v_essence,
            'Gazole': v_gazole,
            'Distance (Mixte)': v_distance
        }
        stacked_data['vehicles'].append(total_v)


        # --- FOOD DETAILS ---
        qs_f = FoodEntry.objects.filter(year=year)
        f_breakdown = {
            'Bœuf': 0, 
            'Porc': 0, 
            'Volaille/Poisson': 0, 
            'Végétarien': 0,
            'Pique-nique': 0
        }
        total_f = 0
        
        for f in qs_f:
            total_f += float(f.total_co2_kg or 0)
            # Recalcul manuel des parts (approximatif si facteurs ont changé mais acceptable)
            f_breakdown['Bœuf'] += f.beef_meals * float(food_factors_map.get('beef', 0))
            f_breakdown['Porc'] += f.pork_meals * float(food_factors_map.get('pork', 0))
            f_breakdown['Volaille/Poisson'] += f.poultry_fish_meals * float(food_factors_map.get('poultry_fish', 0))
            f_breakdown['Végétarien'] += f.vegetarian_meals * float(food_factors_map.get('vegetarian', 0))
            f_breakdown['Pique-nique'] += (f.picnic_meat_meals * float(food_factors_map.get('picnic_meat', 0)) +
                                           f.picnic_no_meat_meals * float(food_factors_map.get('picnic_veg', 0)))

        details_data[year]['food'] = f_breakdown
        stacked_data['food'].append(total_f)


        # --- PURCHASE DETAILS ---
        qs_p = PurchaseData.objects.filter(year=year)
        p_breakdown = {} # Dynamique selon catégories
        total_p = 0
        
        for p in qs_p:
            total_p += float(p.total_co2_kg or 0)
            cat_label = p.get_category_display()
            p_breakdown[cat_label] = p_breakdown.get(cat_label, 0) + float(p.total_co2_kg or 0)
            
        details_data[year]['purchases'] = p_breakdown
        stacked_data['purchases'].append(total_p)


        # --- NUMERIQUE DETAILS ---
        qs_n = EquipementNumerique.objects.filter(year=year)
        n_breakdown = {}
        total_n = 0
        
        for n in qs_n:
            total_n += float(n.total_co2_kg or 0)
            # Friendly Label (remove code)
            label = n.get_type_equipement_display()
            # If the label is "LAPTOP (250kg...)", we want simpler or full is fine?
            # User request: "juste mettre une fois en haut". Labels in doughnut are fine.
            # Maybe simplify "1. Hardware (Terminaux)" -> just use group? No, n.type_equipement is specific code.
            # n.get_type_equipement_display() returns "Ordinateur Portable (250kg CO2e)".
            # Let's simplify it by removing the impact text in parenthesis if needed, but it helps context.
            # Source view had: label.split('(')[0].strip()
            # Let's do that for cleaner chart.
            label_clean = label.split('(')[0].strip() if '(' in label else label
            n_breakdown[label_clean] = n_breakdown.get(label_clean, 0) + float(n.total_co2_kg or 0)
            
        details_data[year]['numerique'] = n_breakdown
        stacked_data['numerique'].append(total_n)

    return JsonResponse({
        'years': sorted_years,
        'stacked_data': stacked_data,
        'details_data': details_data
    })


@login_required
def export_statistics_view(request):
    """
    Export statistics to Excel with Native Charts.
    """
    import datetime
    from django.http import HttpResponse
    from django.db.models import Sum
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference, Series
    
    from apps.vehicles.models import VehicleData
    from apps.purchases.models import PurchaseData
    from apps.alimentation.models import FoodEntry
    from apps.batiment.models import BuildingEnergyData
    from apps.numerique.models import EquipementNumerique
    
    # 1. Fetch Years
    years = set()
    years.update(VehicleData.objects.values_list('year', flat=True))
    years.update(BuildingEnergyData.objects.values_list('year', flat=True))
    years.update(FoodEntry.objects.values_list('year', flat=True))
    years.update(PurchaseData.objects.values_list('year', flat=True))
    years.update(EquipementNumerique.objects.values_list('year', flat=True))
    
    sorted_years = sorted(list(years))
    if not sorted_years:
        sorted_years = [datetime.datetime.now().year]

    # 2. Setup Workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="statistiques_evry.xlsx"'
    
    wb = Workbook()
    
    # --- SHEET 1: SYNTHÈSE (Graphique) ---
    ws = wb.active
    ws.title = "Évolution Annuelle"
    
    # Headers
    headers = ["Année", "Bâtiments", "Véhicules", "Alimentation", "Achats", "Numérique", "Total (kgCO2e)"]
    ws.append(headers)
    
    # Style Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2d6a4f", end_color="2d6a4f", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Data Rows
    for year in sorted_years:
        b_total = BuildingEnergyData.objects.filter(year=year).aggregate(s=Sum('total_co2_kg'))['s'] or 0
        v_total = VehicleData.objects.filter(year=year).aggregate(s=Sum('total_co2_kg'))['s'] or 0
        f_total = FoodEntry.objects.filter(year=year).aggregate(s=Sum('total_co2_kg'))['s'] or 0
        p_total = PurchaseData.objects.filter(year=year).aggregate(s=Sum('total_co2_kg'))['s'] or 0
        n_total = EquipementNumerique.objects.filter(year=year).aggregate(s=Sum('total_co2_kg'))['s'] or 0
        
        g_total = b_total + v_total + f_total + p_total + n_total
        
        ws.append([
            year,
            round(float(b_total), 2),
            round(float(v_total), 2),
            round(float(f_total), 2),
            round(float(p_total), 2),
            round(float(n_total), 2),
            round(float(g_total), 2)
        ])

    # Chart: Stacked Bar
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Évolution des Émissions par Secteur"
    chart.y_axis.title = "kg CO₂e"
    chart.x_axis.title = "Année"
    chart.grouping = "stacked"
    chart.overlap = 100
    
    # Data References
    # Data: Columns B to F (2 to 6)
    # Categories: Column A (Year)
    data = Reference(ws, min_col=2, min_row=1, max_row=len(sorted_years)+1, max_col=6)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_years)+1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "H2")
    
    # --- SHEET 2: DÉTAILS COMPLETS ---
    ws_d = wb.create_sheet(title="Détails Complets")
    ws_d.append(["Année", "Secteur", "Détail/Catégorie", "Impact (kgCO2e)"])
    
    # Style Header Sheet 2
    for cell in ws_d[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Fill Data
    for year in sorted_years:
        # Bâtiments
        for b in BuildingEnergyData.objects.filter(year=year):
            ws_d.append([year, "Bâtiment", b.site_name, round(float(b.total_co2_kg or 0), 2)])
        
        # Véhicules
        for v in VehicleData.objects.filter(year=year):
            ws_d.append([year, "Véhicule", v.service, round(float(v.total_co2_kg or 0), 2)])
            
        # Alimentation
        for f in FoodEntry.objects.filter(year=year):
            ws_d.append([year, "Alimentation", f.service, round(float(f.total_co2_kg or 0), 2)])
            
        # Achats
        for p in PurchaseData.objects.filter(year=year):
            ws_d.append([year, "Achat", p.get_category_display(), round(float(p.total_co2_kg or 0), 2)])
            
        # Numérique
        for n in EquipementNumerique.objects.filter(year=year):
            ws_d.append([year, "Numérique", n.nom, round(float(n.total_co2_kg or 0), 2)])

    # Adjust Column Widths
    for sheet in [ws, ws_d]:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response
